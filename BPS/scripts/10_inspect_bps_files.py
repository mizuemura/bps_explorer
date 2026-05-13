"""
10_inspect_bps_files.py

Inspect raw BPS (Census Building Permits Survey) files before writing parsers.
Run from project root:  python scripts/10_inspect_bps_files.py

Inputs:
    data/raw/bps/state/
    data/raw/bps/county/
    data/raw/bps/cbsa/

Outputs:
    outputs/reports/bps_file_inventory.csv
    outputs/reports/bps_sample_rows.txt
"""

import csv
import io
import os
import re
import sys
from pathlib import Path

# Force UTF-8 output on Windows terminals
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf_8"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Paths ──────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_DIRS = {
    "state":  PROJECT_ROOT / "data" / "raw" / "bps" / "state",
    "county": PROJECT_ROOT / "data" / "raw" / "bps" / "county",
    "cbsa":   PROJECT_ROOT / "data" / "raw" / "bps" / "cbsa",
}
OUT_DIR      = PROJECT_ROOT / "outputs" / "reports"
INVENTORY_CSV = OUT_DIR / "bps_file_inventory.csv"
SAMPLES_TXT   = OUT_DIR / "bps_sample_rows.txt"

OUT_DIR.mkdir(parents=True, exist_ok=True)

TEXT_EXTENSIONS  = {".txt", ".csv", ".tsv", ".dat", ".asc", ".prn"}
EXCEL_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}

SAMPLE_LINES = 10

# ── Year inference ─────────────────────────────────────────────────────────
_YEAR_RE = re.compile(r"(20\d{2}|19\d{2})")

def infer_year(filename: str) -> str:
    m = _YEAR_RE.search(filename)
    return m.group(1) if m else ""

# ── Format detection for text files ───────────────────────────────────────
def detect_format(lines: list[str]) -> tuple[str, str]:
    """
    Returns (format_type, delimiter) where format_type is one of:
    fixed-width | comma-delimited | tab-delimited | pipe-delimited | unknown
    """
    if not lines:
        return "empty", ""

    non_empty = [l for l in lines if l.strip()]
    if not non_empty:
        return "empty", ""

    sample = non_empty[:5]

    # Try csv.Sniffer on joined sample
    joined = "\n".join(sample)
    try:
        dialect = csv.Sniffer().sniff(joined, delimiters=",\t|;")
        delim = dialect.delimiter
        name_map = {",": "comma-delimited", "\t": "tab-delimited",
                    "|": "pipe-delimited", ";": "semicolon-delimited"}
        return name_map.get(delim, f"delimited({repr(delim)})"), delim
    except csv.Error:
        pass

    # Heuristic: fixed-width if lines are consistent length with no common delimiters
    lengths = [len(l.rstrip("\n")) for l in non_empty[:10]]
    has_comma = any("," in l for l in sample)
    has_tab   = any("\t" in l for l in sample)
    if not has_comma and not has_tab:
        length_range = max(lengths) - min(lengths) if lengths else 0
        if length_range <= 5:
            return "fixed-width", ""

    return "unknown", ""

# ── Read raw text lines safely ─────────────────────────────────────────────
def read_text_lines(path: Path, n: int = SAMPLE_LINES) -> tuple[list[str], str]:
    """Returns (lines, encoding_used). Tries utf-8, then latin-1."""
    for enc in ("utf-8", "latin-1"):
        try:
            with open(path, encoding=enc, errors="replace") as f:
                lines = [f.readline() for _ in range(n)]
            return lines, enc
        except Exception:
            continue
    return [], "failed"

# ── Read Excel sheets ──────────────────────────────────────────────────────
def read_excel_info(path: Path) -> list[dict]:
    """Returns list of {sheet_name, nrows, ncols, sample_rows_text}."""
    try:
        import openpyxl
    except ImportError:
        return [{"sheet_name": "?", "nrows": "?", "ncols": "?",
                 "sample_rows_text": "[openpyxl not installed]"}]

    results = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= SAMPLE_LINES:
                    break
                rows.append(row)
            nrows = ws.max_row or "?"
            ncols = ws.max_column or "?"
            sample_text = "\n".join(
                "  " + "\t".join("" if c is None else str(c) for c in row)
                for row in rows
            )
            results.append({
                "sheet_name": sname,
                "nrows": nrows,
                "ncols": ncols,
                "sample_rows_text": sample_text,
            })
        wb.close()
    except Exception as e:
        results.append({"sheet_name": "ERROR", "nrows": "?", "ncols": "?",
                        "sample_rows_text": str(e)})
    return results

# Fallback for older .xls
def read_xls_info(path: Path) -> list[dict]:
    try:
        import xlrd
    except ImportError:
        return [{"sheet_name": "?", "nrows": "?", "ncols": "?",
                 "sample_rows_text": "[xlrd not installed]"}]
    results = []
    try:
        wb = xlrd.open_workbook(path)
        for sname in wb.sheet_names():
            ws = wb.sheet_by_name(sname)
            rows = []
            for i in range(min(SAMPLE_LINES, ws.nrows)):
                rows.append(ws.row_values(i))
            sample_text = "\n".join(
                "  " + "\t".join(str(c) for c in row) for row in rows
            )
            results.append({
                "sheet_name": sname,
                "nrows": ws.nrows,
                "ncols": ws.ncols,
                "sample_rows_text": sample_text,
            })
    except Exception as e:
        results.append({"sheet_name": "ERROR", "nrows": "?", "ncols": "?",
                        "sample_rows_text": str(e)})
    return results

# ── Main ───────────────────────────────────────────────────────────────────
def main():
    inventory_rows = []
    sample_buf = io.StringIO()

    total_files = 0
    skipped     = 0

    for geo_level, folder in INPUT_DIRS.items():
        if not folder.exists():
            print(f"[WARN] Folder not found, skipping: {folder}")
            continue

        files = sorted(folder.rglob("*"))
        files = [f for f in files if f.is_file()]

        if not files:
            print(f"[INFO] No files in {folder}")
            continue

        print(f"\n{'='*60}")
        print(f"  {geo_level.upper()}  ({folder})")
        print(f"{'='*60}")

        for fpath in files:
            total_files += 1
            ext  = fpath.suffix.lower()
            size = fpath.stat().st_size
            year = infer_year(fpath.name)
            rel  = fpath.relative_to(PROJECT_ROOT)

            inv = {
                "path":        str(rel),
                "filename":    fpath.name,
                "extension":   ext,
                "size_bytes":  size,
                "size_kb":     round(size / 1024, 1),
                "geo_level":   geo_level,
                "year":        year,
                "format_type": "",
                "delimiter":   "",
                "sheets":      "",
                "encoding":    "",
                "notes":       "",
            }

            header = f"\n{'─'*60}\n[{geo_level}] {fpath.name}  ({size/1024:.1f} KB)  year={year or '?'}\n{'─'*60}"
            sample_buf.write(header + "\n")
            print(header)

            # ── Text / ASCII ───────────────────────────────────────────
            if ext in TEXT_EXTENSIONS:
                lines, enc = read_text_lines(fpath)
                fmt, delim = detect_format(lines)
                inv["format_type"] = fmt
                inv["delimiter"]   = repr(delim) if delim else ""
                inv["encoding"]    = enc

                block = f"Format: {fmt}  |  Encoding: {enc}\nFirst {SAMPLE_LINES} lines:\n"
                for i, ln in enumerate(lines, 1):
                    block += f"  {i:>3}: {ln.rstrip()}\n"
                sample_buf.write(block)
                print(block, end="")

            # ── Excel ──────────────────────────────────────────────────
            elif ext in EXCEL_EXTENSIONS:
                inv["format_type"] = "excel"
                if ext == ".xls":
                    sheets = read_xls_info(fpath)
                else:
                    sheets = read_excel_info(fpath)

                sheet_names = [s["sheet_name"] for s in sheets]
                inv["sheets"] = "; ".join(str(s) for s in sheet_names)

                block = f"Sheets: {sheet_names}\n"
                for s in sheets:
                    block += (f"  Sheet '{s['sheet_name']}': "
                              f"{s['nrows']} rows × {s['ncols']} cols\n")
                    block += s["sample_rows_text"] + "\n"
                sample_buf.write(block)
                print(block, end="")

            # ── Unknown ────────────────────────────────────────────────
            else:
                inv["format_type"] = "unknown"
                inv["notes"] = "unrecognized extension"
                skipped += 1
                sample_buf.write("  [skipped - unrecognized extension]\n")
                print("  [skipped - unrecognized extension]")

            inventory_rows.append(inv)

    # ── Write inventory CSV ────────────────────────────────────────────────
    fieldnames = ["path", "filename", "extension", "size_bytes", "size_kb",
                  "geo_level", "year", "format_type", "delimiter",
                  "sheets", "encoding", "notes"]

    with open(INVENTORY_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(inventory_rows)

    # ── Write sample rows TXT ─────────────────────────────────────────────
    with open(SAMPLES_TXT, "w", encoding="utf-8") as f:
        f.write(f"BPS FILE SAMPLES — {total_files} files inspected\n")
        f.write("=" * 60 + "\n")
        f.write(sample_buf.getvalue())

    # ── Summary ───────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"DONE - {total_files} files inspected, {skipped} skipped")
    print(f"  Inventory : {INVENTORY_CSV}")
    print(f"  Samples   : {SAMPLES_TXT}")

    if total_files == 0:
        print("\n[!] No files found. Place BPS raw files in:")
        for geo, folder in INPUT_DIRS.items():
            print(f"      {folder}")

    # ── Format breakdown ──────────────────────────────────────────────────
    if inventory_rows:
        from collections import Counter
        fmt_counts = Counter(r["format_type"] for r in inventory_rows)
        yr_counts  = Counter(r["year"] for r in inventory_rows if r["year"])
        geo_counts = Counter(r["geo_level"] for r in inventory_rows)
        print("\nFormat breakdown:")
        for k, v in sorted(fmt_counts.items()):
            print(f"  {k:<22} {v} file(s)")
        print("\nYear coverage:")
        for k in sorted(yr_counts):
            print(f"  {k}: {yr_counts[k]} file(s)")
        print("\nGeography levels:")
        for k, v in sorted(geo_counts.items()):
            print(f"  {k:<12} {v} file(s)")


if __name__ == "__main__":
    main()
